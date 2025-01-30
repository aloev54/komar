import shutil
import pandas as pd
from fuzzywuzzy import process
import openpyxl
import pandas as pd
import datetime
import os
from openpyxl.styles import Border, Side, PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory


def convert_to_date(day, month_str, year, month_cell):
    months = {
        "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
        "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
        "ноября": 11, "декабря": 12
    }
    closest_month, score = process.extractOne(month_str, months.keys())
    month = months.get(closest_month.lower())
    if month:
        date_obj = datetime.date(year, month, day)
        return date_obj.strftime("%d-%m-%Y")
    else:
        raise ValueError(f"Неверное название месяца в ячейке {month_cell}: '{month_str}'")


# def find_duplicates_in_excel(ws):
#     rows_checked = set()  # Множество для хранения уже проверенных строк
#     for row_num in range(2, ws.max_row + 1):  # Начинаем со второй строки, так как первая строка обычно содержит заголовки
#         current_row = ws[row_num]
#         current_data = tuple(cell.value for cell in current_row)  # Преобразуем данные строки в кортеж
#         if current_data in rows_checked:  # Если текущая строка уже проверялась, значит, это дубликат
#             print(f"Найден дубликат в строке {row_num}.")
#         else:
#             rows_checked.add(current_data)  # Добавляем данные текущей строки в множество проверенных строк

def find_duplicates_in_excel(ws):
    rows_checked = {}  # Используем словарь для хранения информации о строках
    for row_num in range(2, ws.max_row + 1):
        current_row = ws[row_num]
        current_data = tuple(cell.value for cell in current_row)
        if current_data in rows_checked:
            # Если текущая строка уже присутствует в словаре, то это дубликат
            print(f"Найден дубликат с строкой {rows_checked[current_data]} в строке {row_num}.")
        else:
            rows_checked[current_data] = row_num  # Сохраняем информацию о текущей строке
        # print(rows_checked)


def compare_dates(ws, row):
    date_format = "%d-%m-%Y"
    cell_f_value = ws['F' + str(row)].value
    cell_i_value = ws['I' + str(row)].value

    try:
        if cell_f_value and cell_i_value:
            date_f = datetime.datetime.strptime(cell_f_value, date_format)
            date_i = datetime.datetime.strptime(cell_i_value, date_format)

            if (date_f - date_i).days > 30:
                # Если условие выполняется, добавляем 1 в ячейку 'AP' текущей строки
                ws['AQ' + str(row)] = 0
            else:
                print(f"В строке {row}: Разница между датами в ячейках F и I превышает 30 дней")
    except ValueError as e:
        print(f"Ошибка в строке {row}: Неправильный формат даты. {e}")


def choose_file():
    Tk().withdraw()
    return askopenfilename(title="Выберите файл назначения")


def choose_directory():
    Tk().withdraw()
    return askdirectory(title="Выберите папку источника")


# Функция для конвертации xls в xlsx
def convert_xls_to_xlsx(source_folder):
    for filename in os.listdir(source_folder):
        if filename.endswith('.xls'):
            file_path = os.path.join(source_folder, filename)
            df = pd.read_excel(file_path)
            new_filename = filename.replace('.xls', '.xlsx')
            new_file_path = os.path.join(source_folder, new_filename)
            df.to_excel(new_file_path, index=False)


# Функции для стилей Excel
def apply_border(ws):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border


def apply_red_fill(cell):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.fill = red_fill


def apply_yellow_fill(cell):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.fill = yellow_fill


# Функция для извлечения и преобразования дат
def extract_date_values(ws, cells_group):
    try:
        day_cell, month_cell, year_cell = cells_group
        day_value = ws[day_cell].value
        month_value = ws[month_cell].value
        year_value = ws[year_cell].value

        if any(v is None for v in [day_value, month_value, year_value]):
            raise ValueError(
                f"Одно или несколько значений пустые: День - {day_cell}: {day_value}, Месяц - {month_cell}: {month_value}, Год - {year_cell}: {year_value}")

        if isinstance(day_value, int) and isinstance(year_value, int):
            if isinstance(month_value, int):
                date_obj = datetime.date(year_value, month_value, day_value)
                return date_obj.strftime("%d-%m-%Y")
            elif isinstance(month_value, str):
                return convert_to_date(day_value, month_value, year_value, month_cell)
            else:
                raise ValueError(f"Неверный тип данных для месяца: {month_cell}: {month_value}")
        else:
            raise ValueError(
                f"Неверный тип данных для дня или года: День - {day_cell}: {day_value}, Год - {year_cell}: {year_value}")

    except ValueError as ve:
        print(ve)
        return None
    except TypeError:
        print(
            f"Неверный тип данных: День - {day_cell}: {day_value}, Месяц - {month_cell}: {month_value}, Год - {year_cell}: {year_value}")
        return None


# Основная функция для копирования ячеек

def copy_cells_to_new_file(source_folder, target_file, cells):
    try:
        convert_xls_to_xlsx(source_folder)

        if os.path.exists(target_file):
            wb_target = openpyxl.load_workbook(target_file)
        else:
            wb_target = openpyxl.Workbook()
        ws_target = wb_target.active
        current_row = ws_target.max_row + 1

        for filename in os.listdir(source_folder):
            if filename.endswith('.xlsx'):
                wb_source = openpyxl.load_workbook(os.path.join(source_folder, filename))
                ws_source = wb_source.active
                current_column = 1

                for cell_group in cells:
                    if len(cell_group) == 1:
                        cell_value = ws_source[cell_group[0]].value
                        if cell_value == None:
                            target_cell = ws_target.cell(row=current_row, column=current_column, value=None)
                            apply_red_fill(target_cell)
                        else:
                            ws_target.cell(row=current_row, column=current_column, value=cell_value)
                        current_column += 1
                    else:
                        date_obj = extract_date_values(ws_source, cell_group)
                        if date_obj:
                            target_cell = ws_target.cell(row=current_row, column=current_column, value=date_obj)
                        else:
                            target_cell = ws_target.cell(row=current_row, column=current_column, value=None)
                            apply_red_fill(target_cell)
                        compare_dates(ws_target, current_row)
                        current_column += 1
                find_duplicates_in_excel(ws_target)
                wb_source.close()
                current_row += 1

        apply_border(ws_target)

        wb_target.save(target_file)
        print("Операция завершена успешно!")

    except Exception as e:
        print(f"Произошла ошибка: {e}")


try:
    # source_folder = choose_directory()  # Выбор папки источника
    # target_file = choose_file()  # Выбор файла назначения
    source_folder = "C:/Dev/excel_copy_13_03/work_excel_1/pa_input"
    target_file = "C:/Dev/excel_copy_13_03/work_excel_1/pa_output/result.xlsx"

    # cells_to_merge = [['AG7', 'AH7', 'AJ7'], ['AD38', 'AF38', 'AH38'],['T68', 'R68', 'P68'],['T75', 'R75', 'P75'],['T91', 'R91', 'P91'], ['AO68', 'AM68', 'AK68'], ['AO100', 'AM100', 'AK100'], ['AK112', 'AM112', 'AO112']]
    cells = [['AL34'], ['AD11'], ['AL34'], ['AG7', 'AH7', 'AJ7'], ['AK14'], ['AD38', 'AF38', 'AH38'], ['AD40'],
             ['AG39'], ['T68', 'R68', 'P68'], ['P70'], ['P71'], ['P72'], ['P73'], ['P74'], ['T75', 'R75', 'P75'],
             ['P86'], ['P87'], ['P88'], ['P89'], ['P90'], ['T91', 'R91', 'P91'], ['P102'], ['P105'], ['P129'],
             ['AO68', 'AM68', 'AK68'], ['AK100', 'AM100', 'AO100'], ['AK102'], ['AK104'], ['AK106'], ['AK108'],
             ['AK110'], ['AK112', 'AM112', 'AO112'], ['AK127'], ['P6'], ['P11'], ['P12'], ['P13'], ['P14'], ['P16'],
             ['C35'], ['P20']]
    copy_cells_to_new_file(source_folder, target_file, cells)
    # find_duplicates_in_excel(target_file)


except Exception as e:
    print(f"Произошла ошибка: {e}")
