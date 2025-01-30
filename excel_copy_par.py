import shutil
import pandas as pd
from fuzzywuzzy import process
import openpyxl
import datetime
import os
from openpyxl.styles import Border, Side, PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory


def find_cell_with_value(ws, value):
    """ Поиск ячейки с заданным значением в листе Excel. """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return (cell.row, cell.column)
    return (None, None)


# def convert_to_date(day, month_str, year, month_cell):
#     months = {
#         "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
#         "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
#         "ноября": 11, "декабря": 12
#     }
#     closest_month, score = process.extractOne(month_str, months.keys())
#     month = months.get(closest_month.lower())
#     if month:
#         date_obj = datetime.date(year, month, day)
#         return date_obj.strftime("%d-%m-%Y")
#     else:
#         raise ValueError(f"Неверное название месяца в ячейке {month_cell}: '{month_str}'")


def find_duplicates_in_excel(ws):
    rows_checked = {}
    for row_num in range(2, ws.max_row + 1):
        current_row = ws[row_num]
        current_data = tuple(cell.value for cell in current_row)
        if current_data in rows_checked:
            print(f"Найден дубликат с строкой {rows_checked[current_data]} в строке {row_num}.")
        else:
            rows_checked[current_data] = row_num


def compare_dates(ws, row):
    date_format = "%d-%m-%Y"
    cell_f_value = ws['F' + str(row)].value
    cell_i_value = ws['I' + str(row)].value
    try:
        if cell_f_value and cell_i_value:
            date_f = datetime.datetime.strptime(cell_f_value, date_format)
            date_i = datetime.datetime.strptime(cell_i_value, date_format)
            if (date_f - date_i).days > 30:
                ws['AQ' + str(row)] = 0
            else:
                print(f"В строке {row}: Разница между датами в ячейках F и I превышает 30 дней")
    except ValueError as e:
        print(f"Ошибка в строке {row}: Неправильный формат даты. {e}")


def choose_file():
    Tk().withdraw()
    return askopenfilename(title="Выберите файл назначения")


def choose_directory():
    Tk().withdraw()
    return askdirectory(title="Выберите папку источника")


def convert_xls_to_xlsx(source_folder):
    for filename in os.listdir(source_folder):
        if filename.endswith('.xls'):
            file_path = os.path.join(source_folder, filename)
            df = pd.read_excel(file_path)
            new_filename = filename.replace('.xls', '.xlsx')
            new_file_path = os.path.join(source_folder, new_filename)
            df.to_excel(new_file_path, index=False)


def apply_border(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border


def apply_red_fill(cell):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.fill = red_fill


def apply_yellow_fill(cell):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.fill = yellow_fill


from fuzzywuzzy import process
import datetime

def extract_date_values(day_value, month_value, year_value):
    """Конвертация значений дня, месяца и года в строку даты."""
    months = {
        "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
        "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
        "ноября": 11, "декабря": 12
    }
    if isinstance(month_value, str):
        # Используем fuzzy matching для нахождения наиболее подходящего месяца
        month_value, score = process.extractOne(month_value, months.keys())
        month_value = months.get(month_value.lower(), None)
    if month_value and all([day_value, year_value]):  # Убедимся, что месяц найден и все части даты присутствуют
        try:
            date_obj = datetime.date(int(year_value), int(month_value), int(day_value))
            return date_obj.strftime("%d-%m-%Y")
        except ValueError as e:
            return f"Ошибка в данных даты: {e}"
    return None



# Основная функция для копирования ячеек с использованием смещений

def copy_cells_to_new_file(source_folder, target_file, cells_with_offsets):
    try:
        convert_xls_to_xlsx(source_folder)
        wb_target = openpyxl.load_workbook(target_file) if os.path.exists(target_file) else openpyxl.Workbook()
        ws_target = wb_target.active
        current_row = ws_target.max_row + 1  # Начинаем с новой строки в документе

        for filename in os.listdir(source_folder):
            if filename.endswith('.xlsx'):
                wb_source = openpyxl.load_workbook(os.path.join(source_folder, filename))
                ws_source = wb_source.active
                base_row, base_col = find_cell_with_value(ws_source, "govno")
                if base_row is None or base_col is None:
                    continue

                current_column = 1
                for cell_group_offsets in cells_with_offsets:
                    if len(cell_group_offsets) == 3:  # Обработка даты
                        date_values = [ws_source.cell(row=base_row + offset[0], column=base_col + offset[1]).value for offset in cell_group_offsets]
                        date_str = extract_date_values(*date_values)
                        if date_str is None or date_str.strip() == '':  # Проверка на пустое значение
                            target_cell = ws_target.cell(row=current_row, column=current_column)
                            target_cell.value = None
                            apply_red_fill(target_cell)
                        else:
                            ws_target.cell(row=current_row, column=current_column).value = date_str
                    else:
                        for offset in cell_group_offsets:
                            source_row = base_row + offset[0]
                            source_col = base_col + offset[1]
                            cell_value = ws_source.cell(row=source_row, column=source_col).value
                            target_cell = ws_target.cell(row=current_row, column=current_column)
                            if cell_value is None or cell_value == '':  # Проверка на пустое значение
                                target_cell.value = None
                                apply_red_fill(target_cell)
                            else:
                                target_cell.value = cell_value
                    current_column += 1

                current_row += 1  # Переходим на следующую строку для новой группы данных из следующего файла
                find_duplicates_in_excel(ws_target)
                # compare_dates(ws_target, current_row - 1)  # Проверяем даты на предыдущей заполненной строке
                wb_source.close()

        apply_border(ws_target)
        wb_target.save(target_file)
        print("Операция завершена успешно!")

    except Exception as e:
        print(f"Произошла ошибка: {e}")


# Пример вызова функции
try:
    source_folder = "C:/Dev/excel_copy_13_03/work_excel_1/pa_input"
    target_file = "C:/Dev/excel_copy_13_03/work_excel_1/pa_output/result.xlsx"
    source_folder = choose_directory()  # Выбор папки источника
    target_file = choose_file()  # Выбор файла назначения
    cell_offsets_corrected = [
        [(32, 36)],
        [(4, 1)],
        [(9, 28)],
        [(32, 36)],
        [(5, 31), (5, 32), (5, 34)],
        [(12, 35)],
        [(36, 28), (36, 30), (36, 32)],
        [(38, 28)],
        [(37, 31)],
        [(66, 18), (66, 16), (66, 14)],
        [(68, 14)],
        [(69, 14)],
        [(70, 14)],
        [(71, 14)],
        [(72, 14)],
        [(73, 18), (73, 16), (73, 14)],
        [(84, 14)],
        [(85, 14)],
        [(86, 14)],
        [(87, 14)],
        [(88, 14)],
        [(89, 18), (89, 16), (89, 14)],
        [(100, 14)],
        [(103, 14)],
        [(127, 14)],
        [(66, 39), (66, 37), (66, 35)],
        [(98, 35), (98, 37), (98, 39)],
        [(100, 35)],
        [(102, 35)],
        [(104, 35)],
        [(106, 35)],
        [(108, 35)],
        [(110, 35), (110, 37), (110, 39)],
        [(4, 14)],
        [(9, 14)],
        [(10, 14)],
        [(11, 14)],
        [(12, 14)],
        [(14, 14)],
        [(33, 1)],
        [(18, 14)]
    ]

    copy_cells_to_new_file(source_folder, target_file, cell_offsets_corrected)
except Exception as e:
    print(f"Произошла ошибка: {e}")
