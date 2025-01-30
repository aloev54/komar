import shutil
import pandas as pd
from fuzzywuzzy import process
import openpyxl
import datetime
import os
from openpyxl.styles import Border, Side, PatternFill
from tkinter import Tk, Label, Button, Text, END, filedialog, ttk
import time

def find_cell_with_value(ws, value):
    """Поиск ячейки с заданным значением в листе Excel."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return (cell.row, cell.column)
    return (None, None)

def find_duplicates_in_excel(ws):
    rows_checked = {}
    duplicates = []
    for row_num in range(2, ws.max_row + 1):
        current_row = ws[row_num]
        current_data = tuple(cell.value for cell in current_row)
        if current_data in rows_checked:
            duplicates.append(f"Найден дубликат со строкой {rows_checked[current_data]} в строке {row_num}.")
        else:
            rows_checked[current_data] = row_num
    return duplicates

def compare_dates(ws, row):
    date_format = "%d-%m-%Y"
    cell_f_value = ws['F' + str(row)].value
    cell_i_value = ws['I' + str(row)].value
    try:
        if cell_f_value and cell_i_value:
            date_f = datetime.datetime.strptime(cell_f_value, date_format)
            date_i = datetime.datetime.strptime(cell_i_value, date_format)
            if (date_f - date_i).days > 30:
                ws['AQ' + str(row)] = 0
            else:
                return f"В строке {row}: Разница между датами в ячейках F и I превышает 30 дней"
    except ValueError as e:
        return f"Ошибка в строке {row}: Неправильный формат даты. {e}"
    return None

def choose_file():
    return filedialog.askopenfilename(title="Выберите файл назначения")

def choose_directory():
    return filedialog.askdirectory(title="Выберите папку источника")

def convert_xls_to_xlsx(source_folder):
    for filename in os.listdir(source_folder):
        if filename.endswith('.xls'):
            file_path = os.path.join(source_folder, filename)
            df = pd.read_excel(file_path)
            new_filename = filename.replace('.xls', '.xlsx')
            new_file_path = os.path.join(source_folder, new_filename)
            df.to_excel(new_file_path, index=False)

def apply_border(ws):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

def apply_red_fill(cell):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell.fill = red_fill

def apply_yellow_fill(cell):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.fill = yellow_fill

def extract_date_values(day_value, month_value, year_value):
    """Конвертация значений дня, месяца и года в строку даты."""
    months = {
        "января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
        "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
        "ноября": 11, "декабря": 12
    }
    if isinstance(month_value, str):
        month_value, score = process.extractOne(month_value, months.keys())
        month_value = months.get(month_value.lower(), None)
    if month_value and all([day_value, year_value]):
        try:
            date_obj = datetime.date(int(year_value), int(month_value), int(day_value))
            return date_obj.strftime("%d-%m-%Y")
        except ValueError as e:
            return f"Ошибка в данных даты: {e}"
    return None

def copy_cells_to_new_file(source_folder, target_file, cells_with_offsets, text_widget, progress_bar, time_label):
    try:
        start_time = time.time()
        convert_xls_to_xlsx(source_folder)
        wb_target = openpyxl.load_workbook(target_file) if os.path.exists(target_file) else openpyxl.Workbook()
        ws_target = wb_target.active
        current_row = ws_target.max_row + 1

        files = [f for f in os.listdir(source_folder) if f.endswith('.xlsx')]
        total_files = len(files)

        for idx, filename in enumerate(files):
            duplicates = []  # Initialize duplicates list for each file
            wb_source = openpyxl.load_workbook(os.path.join(source_folder, filename))
            ws_source = wb_source.active
            base_row, base_col = find_cell_with_value(ws_source, "base")
            if base_row is None or base_col is None:
                continue

            current_column = 1
            for cell_group_offsets in cells_with_offsets:
                if len(cell_group_offsets) == 3:
                    date_values = [ws_source.cell(row=base_row + offset[0], column=base_col + offset[1]).value for offset in cell_group_offsets]
                    date_str = extract_date_values(*date_values)
                    if date_str is None or date_str.strip() == '':
                        target_cell = ws_target.cell(row=current_row, column=current_column)
                        target_cell.value = None
                        apply_red_fill(target_cell)
                    else:
                        ws_target.cell(row=current_row, column=current_column).value = date_str
                else:
                    for offset in cell_group_offsets:
                        source_row = base_row + offset[0]
                        source_col = base_col + offset[1]
                        cell_value = ws_source.cell(row=source_row, column=source_col).value
                        target_cell = ws_target.cell(row=current_row, column=current_column)
                        if cell_value is None or cell_value == '':
                            target_cell.value = None
                            apply_red_fill(target_cell)
                        else:
                            target_cell.value = cell_value
                current_column += 1

            current_row += 1

            # Collect duplicates for the current file
            duplicates = find_duplicates_in_excel(ws_target)
            wb_source.close()

            # Update progress bar
            progress_bar['value'] = ((idx + 1) / total_files) * 100
            text_widget.update_idletasks()

        # After all files are processed, display duplicates from the last file
        for duplicate in duplicates:
            text_widget.insert(END, duplicate + "\n")

        apply_border(ws_target)
        wb_target.save(target_file)
        end_time = time.time()
        time_taken = end_time - start_time
        text_widget.insert(END, "Операция завершена успешно!\n")
        time_label.config(text=f"Время выполнения: {time_taken:.2f} секунд")
        progress_bar['value'] = 100  # Ensure the progress bar is 100% at the end
        text_widget.update_idletasks()

    except Exception as e:
        text_widget.insert(END, f"Произошла ошибка: {e}\n")

def on_execute_button_click(source_folder, target_file, text_widget, progress_bar, time_label):
    text_widget.delete('1.0', END)
    time_label.config(text="Время выполнения: 0.00 секунд")
    cell_offsets_corrected = [
        [(32, 36)],
        [(4, 1)],
        [(9, 28)],
        [(32, 36)],
        [(5, 31), (5, 32), (5, 34)],
        [(12, 35)],
        [(36, 28), (36, 30), (36, 32)],
        [(38, 28)],
        [(37, 31)],
        [(66, 18), (66, 16), (66, 14)],
        [(68, 14)],
        [(69, 14)],
        [(70, 14)],
        [(71, 14)],
        [(72, 14)],
        [(73, 18), (73, 16), (73, 14)],
        [(84, 14)],
        [(85, 14)],
        [(86, 14)],
        [(87, 14)],
        [(88, 14)],
        [(89, 18), (89, 16), (89, 14)],
        [(100, 14)],
        [(103, 14)],
        [(127, 14)],
        [(66, 39), (66, 37), (66, 35)],
        [(98, 35), (98, 37), (98, 39)],
        [(100, 35)],
        [(102, 35)],
        [(104, 35)],
        [(106, 35)],
        [(108, 35)],
        [(110, 35), (110, 37), (110, 39)],
        [(4, 14)],
        [(9, 14)],
        [(10, 14)],
        [(11, 14)],
        [(12, 14)],
        [(14, 14)],
        [(33, 1)],
        [(18, 14)]
    ]
    copy_cells_to_new_file(source_folder, target_file, cell_offsets_corrected, text_widget, progress_bar, time_label)

def create_gui():
    root = Tk()
    root.title("Объединение Excel файлов")

    source_folder = ""
    target_file = ""

    def set_source_folder():
        nonlocal source_folder
        source_folder = choose_directory()
        text_widget.insert(END, f"Исходная папка: {source_folder}\n")

    def set_target_file():
        nonlocal target_file
        target_file = choose_file()
        text_widget.insert(END, f"Целевой файл: {target_file}\n")

    label = Label(root, text="Выберите исходную папку и целевой файл")
    label.pack()

    source_button = Button(root, text="Выбрать исходную папку", command=set_source_folder)
    source_button.pack()

    target_button = Button(root, text="Выбрать целевой файл", command=set_target_file)
    target_button.pack()

    text_widget = Text(root, height=20, width=80)
    text_widget.pack()

    progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
    progress_bar.pack()

    time_label = Label(root, text="Время выполнения: 0.00 секунд")
    time_label.pack()

    execute_button = Button(root, text="Выполнить", command=lambda: on_execute_button_click(source_folder, target_file, text_widget, progress_bar, time_label))
    execute_button.pack()

    root.mainloop()

create_gui()
